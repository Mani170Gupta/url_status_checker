from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
import requests
import os
import io
from concurrent.futures import ThreadPoolExecutor, as_completed

app = Flask(__name__)

# Locked to your WordPress domain. Change this if your domain changes.
CORS(app, resources={r"/*": {"origins": "https://sylvrae.com"}}, supports_credentials=False)

MAX_URLS_PER_REQUEST = 200
REQUEST_TIMEOUT = 6  # seconds per URL
BATCH_CONCURRENCY = 10  # how many URLs to check in parallel at once


def check_urls_batch(urls, max_workers=BATCH_CONCURRENCY):
    """Checks a list of URLs using a limited thread pool, so at most `max_workers`
    requests are in flight at once — fast, but never hammers targets or the server."""
    results = [None] * len(urls)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_index = {
            executor.submit(check_single_url, url): i for i, url in enumerate(urls)
        }
        for future in as_completed(future_to_index):
            index = future_to_index[future]
            results[index] = future.result()
    return [r for r in results if r]


def check_single_url(url):
    """Checks one URL and returns a result dict. Shared by all endpoints."""
    url = url.strip()
    if not url:
        return None
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    try:
        resp = requests.head(
            url,
            allow_redirects=True,
            timeout=REQUEST_TIMEOUT,
            headers={"User-Agent": "Mozilla/5.0 (compatible; StatusChecker/1.0)"},
        )
        # Some servers block HEAD requests (405/403) - retry with GET
        if resp.status_code in (405, 403):
            resp = requests.get(
                url,
                allow_redirects=True,
                timeout=REQUEST_TIMEOUT,
                headers={"User-Agent": "Mozilla/5.0 (compatible; StatusChecker/1.0)"},
            )
        return {
            "url": url,
            "status": resp.status_code,
            "final_url": resp.url,
            "redirected": resp.url != url,
        }
    except requests.exceptions.Timeout:
        return {"url": url, "status": "Timeout", "error": "Request timed out"}
    except requests.exceptions.ConnectionError:
        return {"url": url, "status": "Error", "error": "Connection failed"}
    except Exception as e:
        return {"url": url, "status": "Error", "error": str(e)}


def build_excel_response(results, note=None):
    """Builds an in-memory .xlsx file from a list of result dicts and returns it as a download.
    If `note` is provided, it's written as a highlighted banner row above the table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Status Results"

    if note:
        ws.append([note])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        note_cell = ws.cell(row=1, column=1)
        note_cell.font = Font(bold=True, color="9C0000")
        note_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        ws.append([])  # blank spacer row

    ws.append(["URL", "Status Code", "Final URL", "Redirected"])
    for r in results:
        ws.append([
            r.get("url", ""),
            r.get("status", ""),
            r.get("final_url", ""),
            "Yes" if r.get("redirected") else "No",
        ])
    # Auto-width columns roughly
    for col_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="url_status_results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/", methods=["GET"])
def health_check():
    """Simple route to confirm the service is running."""
    return jsonify({"status": "ok", "message": "URL status checker is live"})


@app.route("/check-status", methods=["POST"])
def check_status():
    """Manual paste flow: accepts a JSON list of URLs, returns JSON results."""
    data = request.get_json(silent=True) or {}
    urls = data.get("urls", [])

    if not isinstance(urls, list) or not urls:
        return jsonify({"error": "Please provide a non-empty list of URLs"}), 400

    total_submitted = len(urls)
    urls = urls[:MAX_URLS_PER_REQUEST]
    results = check_urls_batch(urls)

    truncated = total_submitted > MAX_URLS_PER_REQUEST
    return jsonify({
        "results": results,
        "count": len(results),
        "truncated": truncated,
        "total_submitted": total_submitted,
    })


@app.route("/export-excel", methods=["POST"])
def export_excel():
    """Takes the JSON results already shown on-screen and returns them as a downloadable .xlsx."""
    data = request.get_json(silent=True) or {}
    results = data.get("results", [])
    total_submitted = data.get("total_submitted")
    if not results:
        return jsonify({"error": "No results provided"}), 400

    note = None
    if total_submitted and total_submitted > MAX_URLS_PER_REQUEST:
        note = (
            f"Note: {total_submitted} URLs were submitted. "
            f"Only the first {MAX_URLS_PER_REQUEST} were checked due to processing limits."
        )

    return build_excel_response(results, note=note)


@app.route("/check-excel", methods=["POST"])
def check_excel():
    """Upload flow: accepts an uploaded .xlsx/.csv file of URLs, checks them,
    and returns a downloadable .xlsx file with the results directly."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded. Attach it under the 'file' field."}), 400

    uploaded = request.files["file"]
    if uploaded.filename == "":
        return jsonify({"error": "No file selected"}), 400

    try:
        wb = load_workbook(uploaded)
        ws = wb.active
    except Exception:
        return jsonify({"error": "Could not read the file. Please upload a valid .xlsx file."}), 400

    urls = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        if row[0]:
            urls.append(str(row[0]).strip())

    # Skip a header row if the first cell doesn't look like a URL/domain
    if urls and "." not in urls[0]:
        urls = urls[1:]

    if not urls:
        return jsonify({"error": "No URLs found in the first column of the sheet"}), 400

    total_found = len(urls)
    urls = urls[:MAX_URLS_PER_REQUEST]
    results = check_urls_batch(urls)

    note = None
    if total_found > MAX_URLS_PER_REQUEST:
        note = (
            f"Note: This file contained {total_found} URLs. "
            f"Only the first {MAX_URLS_PER_REQUEST} were checked due to processing limits. "
            f"Please split the remaining URLs into a separate file and re-upload."
        )

    return build_excel_response(results, note=note)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)